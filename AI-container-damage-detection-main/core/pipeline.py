# core/pipeline.py
import uuid
import json
import cv2
import numpy as np
import re
from openpyxl import load_workbook
from pathlib import Path
import time


def _convert_numpy_types(obj):
    """Recursively convert numpy types to Python native types for JSON serialization."""
    if isinstance(obj, dict):
        return {k: _convert_numpy_types(v) for k, v in obj.items()}
    elif isinstance(obj, list):
        return [_convert_numpy_types(item) for item in obj]
    elif isinstance(obj, np.bool_):
        return bool(obj)
    elif isinstance(obj, np.integer):
        return int(obj)
    elif isinstance(obj, np.floating):
        return float(obj)
    elif isinstance(obj, np.ndarray):
        return obj.tolist()
    return obj

from core.utils import load_config, draw_results_on_image, extract_iso6346_from_text
from core.detection import Detector
from core.ocr import OCREngine
from core.paddle_ocr import PaddleOCREngine
from core.llm_validator import LLMValidator
from core.analysis import analyze_damages, grade_container_condition, estimate_repair_cost, determine_discard_status
from core.reporting import generate_reports
from core.bbox_utils import validate_bbox, normalize_bbox, filter_small_bboxes, calculate_iou
from core.inspection_postprocessor import postprocess_damages
from core.segmentation import refine_damage_boundaries
from core.severity_model import assess_damage_severity
from core.damage_visualizer import DamageVisualizer


class InspectionPipeline:
    def __init__(self):
        self.config = load_config()
        self.master_detector = Detector(
            self.config['models']['master_detector'],
            self.config['models']['master_class_names']
        )
        self.damage_detector = Detector(
            self.config['models']['damage_detector'],
            self.config['models']['damage_class_names']
        )
        # Use PaddleOCR if configured, fallback to EasyOCR
        ocr_cfg = self.config.get('ocr', {})
        if ocr_cfg.get('engine', 'paddleocr') == 'paddleocr':
            self.ocr_engine = PaddleOCREngine(self.config)
        else:
            self.ocr_engine = OCREngine()
        self.llm_validator = LLMValidator(self.config)
        self.damage_visualizer = DamageVisualizer(self.config)

        self._location_code_catalog = self._load_location_code_catalog()
        self._location_code_index = self._build_location_code_index(self._location_code_catalog)

    def _canonicalize_damage_class(self, d: dict) -> dict:
        class_names = {
            0: "dent_deformation",
            1: "corrosion_rust",
            2: "frame_bend",
            3: "door_hardware_damage",
            4: "hole_tear_crack",
        }

        def _norm_str(v):
            return str(v or '').strip().lower()

        out = d.copy() if isinstance(d, dict) else {}

        raw_class_id = out.get('class_id', None)
        raw_class_name = _norm_str(out.get('class_name'))
        raw_type = _norm_str(out.get('type'))
        raw_damage_type = _norm_str(out.get('damage_type'))
        raw_desc = _norm_str(out.get('description'))

        class_id = None
        if raw_class_id is not None:
            try:
                class_id = int(raw_class_id)
            except Exception:
                class_id = None

        if class_id is None and raw_class_name:
            cn = raw_class_name
            if cn.startswith(('0_', '1_', '2_', '3_', '4_')):
                cn = cn.split('_', 1)[1]
            if cn in class_names.values():
                class_id = {v: k for k, v in class_names.items()}.get(cn)

        if class_id is None:
            t = raw_type or raw_damage_type
            if t in class_names.values():
                class_id = {v: k for k, v in class_names.items()}.get(t)
            else:
                if any(s in t for s in ('dent', 'deform', 'bulge')) or any(s in raw_desc for s in ('dent', 'deform', 'bulge')):
                    class_id = 0
                elif any(s in t for s in ('rust', 'corros')) or any(s in raw_desc for s in ('rust', 'corros')):
                    class_id = 1
                elif any(s in t for s in ('bend', 'bent', 'frame')) or any(s in raw_desc for s in ('bend', 'bent', 'frame')):
                    class_id = 2
                elif any(s in t for s in ('hinge', 'lock', 'handle', 'hardware')) or any(s in raw_desc for s in ('hinge', 'lock', 'handle', 'hardware')):
                    class_id = 3
                elif any(s in t for s in ('hole', 'tear', 'crack')) or any(s in raw_desc for s in ('hole', 'tear', 'crack')):
                    class_id = 4

        if class_id is None:
            class_id = 0

        class_id = max(0, min(4, int(class_id)))
        class_name = class_names.get(class_id, 'dent_deformation')

        out['class_id'] = class_id
        out['class_name'] = class_name
        out['type'] = class_name
        return out

    def _extract_iso6346_loose(self, raw_text: str) -> str:
        if not raw_text:
            return "N/A"
        text = re.sub(r"[^A-Za-z0-9]", "", str(raw_text)).upper()
        m = re.search(r"[A-Z]{4}[0-9]{7}", text)
        return m.group(0) if m else "N/A"

    def _dedupe_damages_by_iou(self, damages: list, iou_threshold: float = 0.7) -> list:
        if not damages:
            return []
        kept = []
        for d in damages:
            if not isinstance(d, dict):
                continue
            bbox = d.get('bbox')
            if not bbox or not isinstance(bbox, (list, tuple)) or len(bbox) < 4:
                kept.append(d)
                continue
            merged = False
            for k_i, k in enumerate(list(kept)):
                kb = k.get('bbox') if isinstance(k, dict) else None
                if not kb or not isinstance(kb, (list, tuple)) or len(kb) < 4:
                    continue
                try:
                    iou = calculate_iou([int(bbox[0]), int(bbox[1]), int(bbox[2]), int(bbox[3])], [int(kb[0]), int(kb[1]), int(kb[2]), int(kb[3])])
                except Exception:
                    iou = 0.0
                if iou < iou_threshold:
                    continue
                dc = float(d.get('confidence', 0.0) or 0.0)
                kc = float(k.get('confidence', 0.0) or 0.0)
                da = int(max(1, int(bbox[2]) * int(bbox[3])))
                ka = int(max(1, int(kb[2]) * int(kb[3])))
                # Keep the more confident; tie-breaker keep smaller (tighter)
                if (dc > kc) or (dc == kc and da < ka):
                    kept[k_i] = d
                merged = True
                break
            if not merged:
                kept.append(d)
        return kept

    def run(self, image: np.ndarray):
        t0 = time.perf_counter()
        report_id = str(uuid.uuid4())
        output_dir = Path('data/outputs').resolve()
        output_dir.mkdir(exist_ok=True)

        perf_cfg = self.config.get('performance', {}) if isinstance(self.config.get('performance', {}), dict) else {}
        fast_mode = bool(perf_cfg.get('fast_mode', False))
        log_timings = bool(perf_cfg.get('log_timings', False))
        run_metadata_detector = bool(perf_cfg.get('run_metadata_detector', True)) and not fast_mode
        run_llm_ocr_validation = bool(perf_cfg.get('run_llm_ocr_validation', True)) and not fast_mode
        run_yolo_secondary = bool(perf_cfg.get('run_yolo_secondary', True)) and not fast_mode
        run_enhanced_llm = bool(perf_cfg.get('run_enhanced_llm', True)) and not fast_mode
        run_segmentation = bool(perf_cfg.get('run_segmentation', True)) and not fast_mode
        run_severity_model = bool(perf_cfg.get('run_severity_model', True)) and not fast_mode
        generate_reports_enabled = bool(perf_cfg.get('generate_reports', True)) and not fast_mode

        fast_vision_model = perf_cfg.get('fast_vision_model') if fast_mode else None
        fast_max_image_dim = int(perf_cfg.get('fast_max_image_dim', 768)) if fast_mode else None
        jpeg_quality = int(perf_cfg.get('jpeg_quality', 85))
        jpeg_quality = max(30, min(95, jpeg_quality))

        high_recall_second_pass = bool(perf_cfg.get('high_recall_second_pass', False))
        second_pass_min_damages = int(perf_cfg.get('second_pass_min_damages', 3))
        second_pass_max_image_dim = int(perf_cfg.get('second_pass_max_image_dim', 1400))
        second_pass_jpeg_quality = int(perf_cfg.get('second_pass_jpeg_quality', 85))
        second_pass_jpeg_quality = max(30, min(95, second_pass_jpeg_quality))
        second_pass_model = perf_cfg.get('second_pass_model')
        second_pass_trigger_if_container_id_missing = bool(perf_cfg.get('second_pass_trigger_if_container_id_missing', True))
        second_pass_trigger_if_unknown_type = bool(perf_cfg.get('second_pass_trigger_if_unknown_type', True))

        timings = {}

        # Prepare original image dims
        orig_h, orig_w = image.shape[:2]

        # Configurable LLM-only downsampling
        va_cfg = self.config.get('vision_analysis', {})
        ds_cfg = va_cfg.get('downsampling', {}) if isinstance(va_cfg.get('downsampling', {}), dict) else {}
        ds_enable = bool(ds_cfg.get('enable', True))
        max_dim = int(ds_cfg.get('max_image_dim', 1280))
        if fast_mode and fast_max_image_dim:
            max_dim = min(max_dim, int(fast_max_image_dim))
        method = str(ds_cfg.get('method', 'area')).lower()
        interp_map = {
            'area': cv2.INTER_AREA,
            'bilinear': cv2.INTER_LINEAR,
            'bicubic': cv2.INTER_CUBIC,
            'lanczos': cv2.INTER_LANCZOS4
        }
        interp = interp_map.get(method, cv2.INTER_AREA)

        ds_image = image
        scale = 1.0
        if ds_enable and max(orig_h, orig_w) > max_dim:
            scale = max_dim / float(max(orig_h, orig_w))
            ds_w = max(1, int(orig_w * scale))
            ds_h = max(1, int(orig_h * scale))
            ds_image = cv2.resize(image, (ds_w, ds_h), interpolation=interp)
        else:
            ds_w, ds_h = orig_w, orig_h

        # Convert downsampled image to bytes for GPT Vision
        encode_params = [int(cv2.IMWRITE_JPEG_QUALITY), int(jpeg_quality)]
        _, image_buf = cv2.imencode('.jpg', ds_image, encode_params)
        image_bytes = image_buf.tobytes()
        # Also encode original for pricing comparisons (per MB mode or before-DS estimate)
        _, orig_buf = cv2.imencode('.jpg', image, encode_params)
        orig_image_bytes = orig_buf.tobytes()
        image_shape = (ds_h, ds_w)  # shape used by LLM

        timings['encode_ms'] = int(round((time.perf_counter() - t0) * 1000))

        # 1. Detect Metadata ROIs
        t_meta = time.perf_counter()
        metadata_rois = self.master_detector.detect(image) if run_metadata_detector else []
        timings['metadata_ms'] = int(round((time.perf_counter() - t_meta) * 1000))

        # 2. Extract and Validate Container ID with GPT-5 (with vision fallback)
        raw_text = ""
        container_id_roi = next((roi for roi in metadata_rois if roi['class'] == 'container_id'), None)
        if container_id_roi:
            raw_text = self.ocr_engine.extract_text_from_roi(image, container_id_roi['bbox'])

        # Pass image_bytes to enable vision-based container number extraction
        t_ocr = time.perf_counter()
        if run_llm_ocr_validation:
            validated_metadata = self.llm_validator.validate_and_correct_ocr(raw_text, image_bytes)
            container_id = validated_metadata.get("container_number", "N/A")
        else:
            validated_metadata = {}
            container_id = extract_iso6346_from_text(raw_text)
        timings['ocr_validation_ms'] = int(round((time.perf_counter() - t_ocr) * 1000))

        # Normalize/validate container number format (ISO 6346) using local algorithm.
        # Do NOT invent a container number: only accept an ISO 6346-valid ID.
        normalized_from_llm = extract_iso6346_from_text(container_id)
        if normalized_from_llm != "N/A":
            container_id = normalized_from_llm

        # Fallback: derive ISO 6346 ID from OCR text if LLM returned N/A
        # In fast_mode, skip full-frame OCR fallback (too slow) and keep container_id as N/A.
        if (not container_id or container_id == "N/A") and (not fast_mode):
            fallback_id = extract_iso6346_from_text(raw_text)
            if fallback_id == "N/A":
                # Try full-frame OCR as a last resort
                h, w = image.shape[:2]
                full_text = self.ocr_engine.extract_text_from_roi(image, [0, 0, w, h])
                fallback_id = extract_iso6346_from_text(full_text)
            container_id = fallback_id

        # 3. GPT-5 Vision Damage Analysis (Primary) - using downsampled bytes and dims
        va_profile = str(va_cfg.get('damage_prompt_profile', 'yolo') or 'yolo').strip().lower()
        t_vision = time.perf_counter()
        gpt5_analysis = self.llm_validator.analyze_image_damage(
            image_bytes,
            image_shape,
            profile=va_profile,
            model_override=fast_vision_model,
        )
        timings['vision_ms'] = int(round((time.perf_counter() - t_vision) * 1000))

        # Insurance-grade profile: return strict JSON + annotated image (do not mix with YOLO/heuristics)
        if va_profile == 'insurance':
            insurance_result = gpt5_analysis if isinstance(gpt5_analysis, dict) else {}
            damages = insurance_result.get('damages', []) if isinstance(insurance_result.get('damages', []), list) else []

            # Rescale insurance xyxy boxes back to original image coordinates if downsampled
            if scale != 1.0 and damages:
                sx = orig_w / float(ds_w)
                sy = orig_h / float(ds_h)
                for d in damages:
                    if not isinstance(d, dict):
                        continue
                    bb = d.get('bounding_box')
                    if not isinstance(bb, dict):
                        continue
                    try:
                        bb['x_min'] = int(round(int(bb.get('x_min', 0)) * sx))
                        bb['x_max'] = int(round(int(bb.get('x_max', 0)) * sx))
                        bb['y_min'] = int(round(int(bb.get('y_min', 0)) * sy))
                        bb['y_max'] = int(round(int(bb.get('y_max', 0)) * sy))
                    except Exception:
                        continue

            # Normalize/clamp insurance result (xyxy ints inside image)
            img_h, img_w = image.shape[:2]
            normalized_damages = []
            for idx, d in enumerate(damages, start=1):
                if not isinstance(d, dict):
                    continue
                bb = d.get('bounding_box')
                if not isinstance(bb, dict):
                    continue
                try:
                    x1 = int(bb.get('x_min', 0))
                    y1 = int(bb.get('y_min', 0))
                    x2 = int(bb.get('x_max', 0))
                    y2 = int(bb.get('y_max', 0))
                except Exception:
                    continue

                x1 = max(0, min(x1, img_w - 1))
                y1 = max(0, min(y1, img_h - 1))
                x2 = max(0, min(x2, img_w))
                y2 = max(0, min(y2, img_h))
                if x2 <= x1 or y2 <= y1:
                    continue

                damage_id = str(d.get('damage_id') or '').strip() or str(idx)
                damage_type = str(d.get('type') or '').strip() or 'unknown'
                location_code = str(d.get('location_code') or '').strip() or 'UNKNOWN'
                severity = str(d.get('severity') or '').strip() or 'minor'
                validated = bool(d.get('validated', True))

                normalized_damages.append({
                    'damage_id': damage_id,
                    'type': damage_type,
                    'location_code': location_code,
                    'severity': severity,
                    'bounding_box': {
                        'x_min': x1,
                        'y_min': y1,
                        'x_max': x2,
                        'y_max': y2,
                    },
                    'validated': validated,
                })

            for d in normalized_damages:
                if not isinstance(d, dict):
                    continue
                if str(d.get('location_code') or '').strip().upper() in ('', 'UNKNOWN', 'N/A', 'NA'):
                    bb = d.get('bounding_box')
                    if isinstance(bb, dict):
                        try:
                            x1 = int(bb.get('x_min', 0))
                            y1 = int(bb.get('y_min', 0))
                            x2 = int(bb.get('x_max', 0))
                            y2 = int(bb.get('y_max', 0))
                            cx = int((x1 + x2) / 2)
                            cy = int((y1 + y2) / 2)
                        except Exception:
                            continue
                        derived = self._derive_location_code(
                            bbox_center=(cx, cy),
                            image_shape=image.shape[:2],
                            damage_type=str(d.get('type') or ''),
                            zone=str(d.get('zone') or ''),
                        )
                        if derived:
                            d['location_code'] = derived

            # Prefer OCR-derived container_id if insurance model didn't produce a valid one
            insurance_container_id = insurance_result.get('container_id')
            insurance_container_id = str(insurance_container_id).strip() if insurance_container_id is not None else ''
            if not insurance_container_id or insurance_container_id.lower() in ('n/a', 'na', 'unknown'):
                insurance_container_id = container_id

            try:
                insp_conf = insurance_result.get('inspection_confidence')
                insp_conf = int(float(str(insp_conf).strip()))
            except Exception:
                insp_conf = 0
            insp_conf = max(0, min(100, insp_conf))

            insurance_report = {
                'container_id': insurance_container_id,
                'inspection_confidence': insp_conf,
                'damages': normalized_damages,
            }

            insurance_dir = output_dir / 'insurance_previews'
            insurance_dir.mkdir(parents=True, exist_ok=True)
            insurance_img_path = str(insurance_dir / f"{report_id}_annotated.jpg")
            try:
                annotated = self.damage_visualizer.create_insurance_annotated_image(image, normalized_damages)
                cv2.imwrite(insurance_img_path, annotated)
            except Exception as e:
                print(f"WARNING: Insurance visualization generation failed: {e}")

            report = {
                'report_id': report_id,
                **insurance_report,
                'artifacts': {
                    'annotated_image': insurance_img_path,
                    'insurance_annotated_image': insurance_img_path,
                },
            }
            if log_timings:
                timings['total_ms'] = int(round((time.perf_counter() - t0) * 1000))
                report['timings'] = timings
            report = _convert_numpy_types(report)

            json_path = str(output_dir / f"{report_id}.json")
            with open(json_path, 'w') as f:
                json.dump(report, f, indent=2)

            return report

        gpt5_damages = gpt5_analysis.get('damages', [])
        container_details = gpt5_analysis.get('container_details', {})

        if isinstance(gpt5_damages, list):
            gpt5_damages = [self._canonicalize_damage_class(d) for d in gpt5_damages if isinstance(d, dict)]
        else:
            gpt5_damages = []

        # Rescale GPT bboxes back to original image coordinates if downsampled
        if scale != 1.0 and gpt5_damages:
            sx = orig_w / float(ds_w)
            sy = orig_h / float(ds_h)
            for d in gpt5_damages:
                if 'bbox' in d and d['bbox'] and len(d['bbox']) == 4:
                    x, y, w, h = d['bbox']
                    d['bbox'] = [int(round(x * sx)), int(round(y * sy)), int(round(w * sx)), int(round(h * sy))]
        if scale != 1.0 and container_details:
            sx = orig_w / float(ds_w)
            sy = orig_h / float(ds_h)
            if container_details.get('container_number_bbox'):
                x, y, w, h = container_details['container_number_bbox']
                container_details['container_number_bbox'] = [int(round(x * sx)), int(round(y * sy)), int(round(w * sx)), int(round(h * sy))]
            if isinstance(container_details.get('data_plate_info'), dict) and container_details['data_plate_info'].get('bbox'):
                x, y, w, h = container_details['data_plate_info']['bbox']
                container_details['data_plate_info']['bbox'] = [int(round(x * sx)), int(round(y * sy)), int(round(w * sx)), int(round(h * sy))]
            if container_details.get('location_codes'):
                for loc in container_details['location_codes']:
                    if isinstance(loc, dict) and loc.get('bbox'):
                        x, y, w, h = loc['bbox']
                        loc['bbox'] = [int(round(x * sx)), int(round(y * sy)), int(round(w * sx)), int(round(h * sy))]

        need_second_pass = False
        if va_profile != 'insurance' and high_recall_second_pass:
            if len(gpt5_damages) < max(0, second_pass_min_damages):
                need_second_pass = True
            if second_pass_trigger_if_unknown_type and any(str(d.get('type') or '').strip().lower() in ('', 'unknown', 'n/a') for d in gpt5_damages):
                need_second_pass = True
            if second_pass_trigger_if_container_id_missing and (not container_id or container_id == 'N/A'):
                need_second_pass = True

        if need_second_pass:
            t_vision2 = time.perf_counter()
            max_dim2 = int(second_pass_max_image_dim or max_dim)
            max_dim2 = max(256, max_dim2)
            ds2 = image
            oh, ow = image.shape[:2]
            if max(oh, ow) > max_dim2:
                scale2 = max_dim2 / float(max(oh, ow))
                dw2 = max(1, int(ow * scale2))
                dh2 = max(1, int(oh * scale2))
                ds2 = cv2.resize(image, (dw2, dh2), interpolation=interp)
            else:
                dw2, dh2 = ow, oh

            encode_params2 = [int(cv2.IMWRITE_JPEG_QUALITY), int(second_pass_jpeg_quality)]
            _, buf2 = cv2.imencode('.jpg', ds2, encode_params2)
            image_bytes2 = buf2.tobytes()
            shape2 = (dh2, dw2)

            gpt5_analysis2 = self.llm_validator.analyze_image_damage(
                image_bytes2,
                shape2,
                profile=va_profile,
                model_override=second_pass_model,
            )
            timings['vision_second_pass_ms'] = int(round((time.perf_counter() - t_vision2) * 1000))

            damages2 = gpt5_analysis2.get('damages', []) if isinstance(gpt5_analysis2, dict) else []
            details2 = gpt5_analysis2.get('container_details', {}) if isinstance(gpt5_analysis2, dict) else {}

            if isinstance(damages2, list) and damages2:
                sx2 = ow / float(dw2)
                sy2 = oh / float(dh2)
                norm2 = []
                for d in damages2:
                    if not isinstance(d, dict):
                        continue
                    dd = self._canonicalize_damage_class(d)
                    if dd.get('bbox') and isinstance(dd.get('bbox'), (list, tuple)) and len(dd.get('bbox')) == 4:
                        try:
                            x, y, w, h = dd['bbox']
                            dd['bbox'] = [int(round(x * sx2)), int(round(y * sy2)), int(round(w * sx2)), int(round(h * sy2))]
                        except Exception:
                            pass
                    norm2.append(dd)
                gpt5_damages.extend(norm2)

            if isinstance(details2, dict) and details2 and (not container_details):
                container_details = details2
            if isinstance(details2, dict) and details2 and (isinstance(container_details, dict)):
                if not container_details.get('container_number') and details2.get('container_number'):
                    container_details['container_number'] = details2.get('container_number')

        # 4. Traditional YOLO Detection (Secondary/Validation)
        if run_yolo_secondary:
            yolo_damages = self.damage_detector.detect(image, self.config['models']['detection_confidence_threshold'])
        else:
            yolo_damages = []

        # 5. Combine and validate detections
        combined_damages = self._combine_damage_detections(gpt5_damages, yolo_damages)

        # 5a. Validate and normalize bounding boxes (use original image shape)
        image_shape = image.shape[:2]  # (height, width)
        combined_damages = self._validate_and_normalize_bboxes(combined_damages, image_shape)

        # De-duplicate near-identical boxes (e.g., second-pass + first-pass duplicates)
        combined_damages = self._dedupe_damages_by_iou(combined_damages, iou_threshold=0.72)

        # 5b. Validate container details bounding boxes
        if container_details:
            container_details = self._validate_container_details_bboxes(container_details, image_shape)

        if (not container_id or container_id == "N/A") and isinstance(container_details, dict):
            cid_from_vision = extract_iso6346_from_text(container_details.get('container_number'))
            if cid_from_vision != "N/A":
                container_id = cid_from_vision

        # 6. Enhanced GPT Vision analysis (uses downsampled image); no extra scaling needed as we rescale later
        if run_enhanced_llm:
            enhanced_analysis = self.llm_validator.enhanced_damage_analysis(image_bytes, combined_damages)
        else:
            enhanced_analysis = {}
        # Rescale any enhanced bbox back to original
        if scale != 1.0 and isinstance(enhanced_analysis, dict) and enhanced_analysis.get('enhanced_damages'):
            sx = orig_w / float(ds_w)
            sy = orig_h / float(ds_h)
            for d in enhanced_analysis['enhanced_damages']:
                if 'bbox' in d and d['bbox'] and len(d['bbox']) == 4:
                    x, y, w, h = d['bbox']
                    d['bbox'] = [int(round(x * sx)), int(round(y * sy)), int(round(w * sx)), int(round(h * sy))]

        # 7. Quality Control Gate with GPT-5 confidence (improved logic)
        gpt5_confidence = gpt5_analysis.get('confidence_score', 0.0)

        # Smart confidence calculation: prioritize GPT-5 vision for damage analysis
        # Only penalize if metadata detection completely failed AND GPT-5 confidence is low
        if metadata_rois:
            metadata_confidence = np.mean([roi['confidence'] for roi in metadata_rois])
            # Weighted average: GPT-5 (70%) + Metadata (30%) for damage assessment
            final_confidence = (gpt5_confidence * 0.7) + (metadata_confidence * 0.3)
        else:
            # No metadata found - rely entirely on GPT-5 vision for damage assessment
            final_confidence = gpt5_confidence

        # Only flag for manual review if confidence is truly low AND no damages detected
        has_damages = len(combined_damages) > 0 or gpt5_analysis.get('overall_condition') not in ['Excellent', 'Good', 'Unknown']

        if final_confidence < self.config['models']['high_confidence_threshold'] and not has_damages:
            return {
                "status": "Unclear – Needs Manual Review",
                "confidence": final_confidence,
                "gpt5_analysis": gpt5_analysis,
                "enhanced_analysis": enhanced_analysis
            }

        # 8. Process damages with enhanced data
        processed_damages = self._process_enhanced_damages(combined_damages, enhanced_analysis)

        # Ensure canonical 5-class types survive enhanced processing
        processed_damages = [self._canonicalize_damage_class(d) if isinstance(d, dict) else d for d in (processed_damages or [])]

        # 8a. SAM/Mask R-CNN segmentation for precise damage boundaries
        if run_segmentation:
            try:
                processed_damages = refine_damage_boundaries(image, processed_damages, self.config)
            except Exception as e:
                print(f"WARNING: Segmentation refinement failed: {e}")

        # 8b. Custom ML severity model for accurate severity scoring
        if run_severity_model:
            try:
                processed_damages = assess_damage_severity(image, processed_damages, self.config)
            except Exception as e:
                print(f"WARNING: ML severity assessment failed: {e}")

        # 8c. Postprocess detections (no hallucinated boxes): refine/split overlaps, classify, score severity,
        # flag false positives, and conservatively mark items for human review when confidence < 0.85.
        post_cfg = self.config.get('inspection_postprocess', {}) if isinstance(self.config.get('inspection_postprocess', {}), dict) else {}
        min_confidence_for_auto = float(post_cfg.get('min_confidence_for_auto', 0.85))
        postprocessed = postprocess_damages(
            processed_damages,
            image_shape=image.shape[:2],
            min_confidence_for_auto=min_confidence_for_auto,
        )
        processed_damages = postprocessed.get('damages', processed_damages) if isinstance(postprocessed, dict) else processed_damages
        postprocess_summary = postprocessed.get('summary', {}) if isinstance(postprocessed, dict) else {}

        # Canonicalize again after postprocess (UI reads `type`, and postprocess adds `damage_type`)
        processed_damages = [self._canonicalize_damage_class(d) if isinstance(d, dict) else d for d in (processed_damages or [])]

        damages_with_costs, total_cost = estimate_repair_cost(processed_damages, self.config)

        # 9. Grade Overall Condition using GPT-5 assessment
        condition = gpt5_analysis.get('overall_condition', grade_container_condition(damages_with_costs, self.config))

        # 9a. Determine discard status
        discard_status = determine_discard_status(damages_with_costs, condition)

        # 9b. Add location codes to damages
        damages_with_location_codes = self._add_location_codes_to_damages(damages_with_costs, container_details, image.shape[:2])

        # Extract owner from container details or metadata
        container_owner = validated_metadata.get("owner", "N/A")
        if container_owner == "N/A" and container_details:
            container_owner = container_details.get("owner_code", "N/A")

        # If strict ISO 6346 validation failed, allow best-effort candidate (marked as unvalidated)
        container_id_checkdigit_valid = (container_id not in (None, "", "N/A"))
        if container_id in (None, "", "N/A") and isinstance(container_details, dict):
            loose = self._extract_iso6346_loose(container_details.get('container_number'))
            if loose != "N/A":
                container_id = loose
                container_id_checkdigit_valid = False

        # Extract container type from container details or metadata
        container_type = validated_metadata.get("container_type", "N/A")
        if container_type == "N/A" and container_details:
            container_type = container_details.get("container_type", "N/A")

        # Compute cost estimates (input image cost dominated by vision API); estimate based on megapixels and calls
        pricing = self.config.get('llm_pricing', {})
        pricing_mode = str(pricing.get('pricing_mode', 'per_mp')).lower()
        price_per_mp_usd = float(pricing.get('price_per_megapixel_usd', 0.004))
        price_per_mb_usd = float(pricing.get('price_per_mb_usd', 0.060))
        calls_per_image = int(pricing.get('calls_per_image', 3))
        usd_to_inr = float(self.config.get('currency', {}).get('usd_to_inr', 83.0))

        orig_mp = (orig_w * orig_h) / 1e6
        down_mp = (ds_w * ds_h) / 1e6
        if pricing_mode == 'per_mb':
            orig_mb = max(1e-6, len(orig_image_bytes) / (1024.0 * 1024.0))
            down_mb = max(1e-6, len(image_bytes) / (1024.0 * 1024.0))
            cost_before_usd = orig_mb * price_per_mb_usd * calls_per_image
            cost_after_usd = down_mb * price_per_mb_usd * calls_per_image
        else:
            cost_before_usd = orig_mp * price_per_mp_usd * calls_per_image
            cost_after_usd = down_mp * price_per_mp_usd * calls_per_image
        cost_estimate = {
            "calls_per_image": calls_per_image,
            "original_resolution": {"width": orig_w, "height": orig_h, "megapixels": round(orig_mp, 3)},
            "downsampled_resolution": {"width": ds_w, "height": ds_h, "megapixels": round(down_mp, 3)},
            "pricing_mode": pricing_mode,
            "price_per_megapixel_usd": price_per_mp_usd,
            "price_per_mb_usd": price_per_mb_usd,
            "estimated_cost_before_usd": round(cost_before_usd, 6),
            "estimated_cost_after_usd": round(cost_after_usd, 6),
            "estimated_cost_before_inr": round(cost_before_usd * usd_to_inr, 4),
            "estimated_cost_after_inr": round(cost_after_usd * usd_to_inr, 4)
        }

        total_ai_cost_usd = round(cost_after_usd, 6)
        total_ai_cost_inr = round(cost_after_usd * usd_to_inr, 4)

        # 10. Assemble Enhanced Report Data
        report_data = {
            "report_id": report_id,
            "container_id": container_id,
            "container_id_checkdigit_valid": bool(container_id_checkdigit_valid),
            "container_type": container_type,
            "container_owner": container_owner,
            "size_type_code": container_details.get("size_type_code", "N/A") if container_details else "N/A",
            "equipment_category": container_details.get("equipment_category", "N/A") if container_details else "N/A",
            "condition": condition,
            "estimated_cost": total_cost,
            "currency": "INR",
            "detected_damages": damages_with_location_codes,
            "confidence": final_confidence,
            "gpt5_analysis": gpt5_analysis,
            "enhanced_analysis": enhanced_analysis,
            "structural_integrity": gpt5_analysis.get('structural_integrity', 'Unknown'),
            "safety_concerns": gpt5_analysis.get('safety_concerns', []),
            "maintenance_recommendations": gpt5_analysis.get('maintenance_recommendations', []),
            "container_details": container_details,
            "discard_status": discard_status,
            "cost_estimate": cost_estimate,
            "total_ai_cost_usd": total_ai_cost_usd,
            "total_ai_cost_inr": total_ai_cost_inr,
            "needs_human_review": (final_confidence < min_confidence_for_auto) or bool(postprocess_summary.get('needs_human_review_count', 0)),
            "human_review_reasons": [
                "confidence_below_threshold" if final_confidence < min_confidence_for_auto else None,
                "low_confidence_damage_items" if bool(postprocess_summary.get('needs_human_review_count', 0)) else None,
                "false_positive_candidates" if bool(postprocess_summary.get('false_positive_likely_count', 0)) else None
            ],
            "structured_inspection": {
                "inspection_standard": "IICL-style",
                "rules": {
                    "never_hallucinate_damage_or_boxes": True,
                    "use_only_provided_detections_and_ocr": True,
                    "be_conservative_below_confidence": min_confidence_for_auto
                },
                "damage_postprocess_summary": postprocess_summary,
                "ocr": {
                    "raw_text": raw_text,
                    "llm_container_number": validated_metadata.get('container_number') if isinstance(validated_metadata, dict) else None,
                    "final_container_number": container_id,
                    "iso6346_valid": bool(container_id_checkdigit_valid)
                }
            }
        }

        report_data["human_review_reasons"] = [r for r in report_data.get("human_review_reasons", []) if r]

        # 11. Generate Visuals and Reports
        # 11a. Save legacy annotated image (no ticks/labels; includes metadata boxes)
        legacy_annotated_image = draw_results_on_image(image, damages_with_costs, metadata_rois, container_id, container_details)
        legacy_img_path = str(output_dir / f"{report_id}_annotated_legacy.jpg")
        cv2.imwrite(legacy_img_path, legacy_annotated_image)

        # 11b. Generate reference-style visualization (damage boxes + ticks + labels + short description)
        img_path = legacy_img_path
        yolo_output_config = self.config.get('yolo_output', {}) if isinstance(self.config.get('yolo_output', {}), dict) else {}
        if yolo_output_config.get('generate_qa_preview', True):
            try:
                viz_damages = self._prepare_visualization_damages(damages_with_location_codes, image.shape[:2])
                viz_damages = self._filter_yolo_training_outputs(viz_damages, image.shape[:2])
                damage_visualization_data = {
                    'damages': viz_damages,
                    'damage_summary': {
                        'total_damages': len(viz_damages),
                        'class_counts': self._calculate_class_counts(viz_damages)
                    }
                }

                yolo_viz_dir = output_dir / yolo_output_config.get('output_structure', {}).get('qa_preview_dir', 'qa_previews')
                yolo_viz_dir.mkdir(parents=True, exist_ok=True)

                yolo_annotated_path = yolo_viz_dir / f"{report_id}_annotated.jpg"
                yolo_annotated_image = self.damage_visualizer.create_annotated_image(image, viz_damages)
                cv2.imwrite(str(yolo_annotated_path), yolo_annotated_image)

                # Make the YOLO preview the primary image for UI + downloads
                img_path = str(yolo_annotated_path)
                report_data['yolo_annotated_image'] = str(yolo_annotated_path)
                report_data['legacy_annotated_image'] = legacy_img_path

                if yolo_output_config.get('generate_labels', True):
                    yolo_labels_dir = output_dir / yolo_output_config.get('output_structure', {}).get('labels_dir', 'labels')
                    yolo_labels_dir.mkdir(parents=True, exist_ok=True)
                    yolo_label_path = yolo_labels_dir / f"{report_id}.txt"
                    self.damage_visualizer.save_yolo_labels(
                        damage_visualization_data,
                        str(yolo_label_path),
                        orig_w,
                        orig_h,
                    )
                    report_data['yolo_label_file'] = str(yolo_label_path)

            except Exception as e:
                print(f"WARNING: YOLO visualization generation failed: {e}")

        report_files = generate_reports(report_data, img_path, output_dir) if generate_reports_enabled else {}

        report_data['artifacts'] = {
            "annotated_image": img_path,
            "yolo_annotated_image": report_data.get("yolo_annotated_image"),
            "legacy_annotated_image": report_data.get("legacy_annotated_image"),
            "yolo_label_file": report_data.get("yolo_label_file"),
            **report_files,
        }

        if log_timings:
            timings['total_ms'] = int(round((time.perf_counter() - t0) * 1000))
            report_data['timings'] = timings

        # Convert numpy types to native Python types for JSON serialization
        report_data = _convert_numpy_types(report_data)

        json_path = str(output_dir / f"{report_id}.json")
        with open(json_path, 'w') as f:
            json.dump(report_data, f, indent=2)

        return report_data

    def _filter_yolo_training_outputs(self, damages: list, image_shape: tuple) -> list:
        """Filter out oversized / likely-false-positive boxes from YOLO preview and labels."""
        yolo_cfg = self.config.get('yolo_output', {}) if isinstance(self.config.get('yolo_output', {}), dict) else {}
        enable = bool(yolo_cfg.get('filter_false_positives_in_preview', False))
        if not enable:
            return damages

        img_h, img_w = image_shape[:2]
        img_area = float(max(1, img_h * img_w))

        max_area_ratio = float(yolo_cfg.get('max_damage_box_area_ratio', 0.25))
        max_wh_ratio = float(yolo_cfg.get('max_damage_box_wh_ratio', 0.60))

        kept = []
        for d in damages or []:
            if not isinstance(d, dict):
                continue
            bbox = d.get('bbox')
            if not isinstance(bbox, (list, tuple)) or len(bbox) < 4:
                continue
            try:
                x, y, w, h = int(bbox[0]), int(bbox[1]), int(bbox[2]), int(bbox[3])
            except Exception:
                continue
            if w <= 0 or h <= 0:
                continue

            # Drop explicit false positives (from postprocess)
            if bool(d.get('false_positive_likely')):
                continue

            # Drop boxes that are too large for a single localized defect
            area_ratio = (float(w * h) / img_area)
            if area_ratio > max_area_ratio:
                continue

            if (w / float(max(1, img_w)) > max_wh_ratio) and (h / float(max(1, img_h)) > max_wh_ratio):
                continue

            kept.append(d)

        return kept

    def _prepare_visualization_damages(self, damages: list, image_shape: tuple) -> list:
        img_h, img_w = image_shape[:2]
        class_names = {
            0: "dent_deformation",
            1: "corrosion_rust",
            2: "frame_bend",
            3: "door_hardware_damage",
            4: "hole_tear_crack",
        }

        type_to_class = {
            'dent': 0,
            'dent_deformation': 0,
            'deformation': 0,
            'rust': 1,
            'corrosion': 1,
            'corrosion_rust': 1,
            'bent_frame': 2,
            'frame_bend': 2,
            'door': 3,
            'door_hardware_damage': 3,
            'hole': 4,
            'crack': 4,
            'hole_tear_crack': 4,
            'tear': 4,
        }

        prepared = []
        for d in damages:
            if not isinstance(d, dict):
                continue
            bbox = d.get('bbox')
            if not isinstance(bbox, (list, tuple)) or len(bbox) < 4:
                continue
            x, y, w, h = bbox[0], bbox[1], bbox[2], bbox[3]
            try:
                x, y, w, h = int(x), int(y), int(w), int(h)
            except Exception:
                continue
            if w <= 0 or h <= 0:
                continue

            # Heuristic: some LLM outputs may return xyxy instead of xywh.
            # Convert ONLY when xywh is clearly invalid (goes outside image),
            # but bbox[2]/bbox[3] look like bottom-right coordinates within image.
            xywh_out_of_bounds = (x + w > img_w) or (y + h > img_h) or (w > img_w) or (h > img_h)
            looks_like_xyxy = (w > x) and (h > y) and (w <= img_w) and (h <= img_h)
            if xywh_out_of_bounds and looks_like_xyxy:
                x2, y2 = w, h
                w = max(1, x2 - x)
                h = max(1, y2 - y)

            # Clip to image bounds
            x = max(0, min(x, img_w - 1))
            y = max(0, min(y, img_h - 1))
            w = max(1, min(w, img_w - x))
            h = max(1, min(h, img_h - y))

            class_id = d.get('class_id')
            if class_id is None:
                damage_type = str(d.get('type', 'unknown')).lower()
                class_id = type_to_class.get(damage_type, 0)
            try:
                class_id = int(class_id)
            except Exception:
                class_id = 0
            class_id = max(0, min(4, class_id))

            class_name = d.get('class_name')
            if not class_name:
                class_name = class_names.get(class_id, "dent_deformation")

            description = d.get('description')
            if not description:
                description = f"{class_name}"

            entry = d.copy()
            entry['bbox'] = [x, y, w, h]
            entry['class_id'] = class_id
            entry['class_name'] = class_name
            entry['description'] = description
            prepared.append(entry)

        return prepared

    def _combine_damage_detections(self, gpt5_damages: list, yolo_damages: list) -> list:
        """Combine GPT-5 and YOLO damage detections for maximum accuracy"""
        combined = []

        # Add AI Vision damages (primary source)
        for damage in gpt5_damages:
            if isinstance(damage, dict):
                damage = self._canonicalize_damage_class(damage)
            damage_entry = {
                'type': damage.get('type', 'dent_deformation'),
                'severity': damage.get('severity', 'minor'),
                'zone': damage.get('zone', 'unknown'),
                'confidence': damage.get('confidence', 0.5),
                'description': damage.get('description', ''),
                'repair_priority': damage.get('repair_priority', 'medium'),
                'source': 'ai_vision',
                'class_id': damage.get('class_id'),
                'class_name': damage.get('class_name'),
            }
            # Add bounding box if available
            if 'bbox' in damage and damage['bbox']:
                damage_entry['bbox'] = damage['bbox']
            combined.append(damage_entry)

        # Add YOLO damages that don't overlap with GPT-5
        for damage in yolo_damages:
            # Check for overlap with existing GPT-5 damages
            overlaps = any(
                'bbox' in gpt_damage and gpt_damage.get('bbox') and
                abs(damage['bbox'][0] - gpt_damage['bbox'][0]) < 50 and
                abs(damage['bbox'][1] - gpt_damage['bbox'][1]) < 50
                for gpt_damage in combined
            )

            if not overlaps:
                dd = self._canonicalize_damage_class({
                    'type': damage.get('class', 'unknown'),
                    'description': f"YOLO detected {damage.get('class', 'damage')}"
                })
                combined.append({
                    'type': dd.get('type', 'dent_deformation'),
                    'severity': 'minor',  # Default for YOLO
                    'zone': 'unknown',
                    'confidence': damage.get('confidence', 0.5),
                    'description': dd.get('description', f"YOLO detected {damage.get('class', 'damage')}"),
                    'repair_priority': 'medium',
                    'source': 'yolo',
                    'bbox': damage.get('bbox', []),
                    'class_id': dd.get('class_id'),
                    'class_name': dd.get('class_name'),
                })

        return combined

    def _process_enhanced_damages(self, damages: list, enhanced_analysis: dict) -> list:
        """Process damages with enhanced GPT-5 analysis data"""
        enhanced_damages = enhanced_analysis.get('enhanced_damages', damages)
        repair_recommendations = enhanced_analysis.get('repair_recommendations', [])
        cost_breakdown = enhanced_analysis.get('cost_breakdown', {})

        processed = []
        for i, damage in enumerate(damages):
            enhanced_damage = enhanced_damages[i] if i < len(enhanced_damages) else damage

            processed_damage = {
                'type': enhanced_damage.get('type', damage.get('type', 'unknown')),
                'severity': enhanced_damage.get('severity', damage.get('severity', 'minor')),
                'zone': enhanced_damage.get('zone', damage.get('zone', 'unknown')),
                'confidence': enhanced_damage.get('confidence', damage.get('confidence', 0.5)),
                'description': enhanced_damage.get('description', damage.get('description', '')),
                'repair_priority': enhanced_damage.get('repair_priority', damage.get('repair_priority', 'medium')),
                'source': damage.get('source', 'combined'),
                'repair_recommendation': repair_recommendations[i] if i < len(repair_recommendations) else 'Standard repair',
                'detailed_cost': cost_breakdown.get(damage.get('type', 'unknown'), {})
            }

            if 'bbox' in damage:
                processed_damage['bbox'] = damage['bbox']

            processed.append(processed_damage)

        return processed
    
    def _validate_and_normalize_bboxes(self, damages: list, image_shape: tuple) -> list:
        """Validate and normalize bounding boxes for damages"""
        validated_damages = []
        
        for damage in damages:
            if 'bbox' in damage and damage['bbox']:
                # Normalize bbox to ensure it's within image bounds
                damage['bbox'] = normalize_bbox(damage['bbox'], image_shape)
                
                # Validate bbox
                if validate_bbox(damage['bbox'], image_shape):
                    validated_damages.append(damage)
                else:
                    # Keep damage but remove invalid bbox
                    damage_copy = damage.copy()
                    damage_copy.pop('bbox', None)
                    validated_damages.append(damage_copy)
            else:
                validated_damages.append(damage)
        
        post_cfg = self.config.get('inspection_postprocess', {}) if isinstance(self.config.get('inspection_postprocess', {}), dict) else {}
        min_bbox_area = int(post_cfg.get('min_bbox_area', 50))
        min_bbox_area = max(0, min_bbox_area)
        validated_damages = filter_small_bboxes(validated_damages, min_area=min_bbox_area)
        
        return validated_damages
    
    def _validate_container_details_bboxes(self, container_details: dict, image_shape: tuple) -> dict:
        """Validate and normalize container details bounding boxes"""
        validated_details = container_details.copy()
        
        # Validate container number bbox
        if 'container_number_bbox' in validated_details and validated_details['container_number_bbox']:
            bbox = normalize_bbox(validated_details['container_number_bbox'], image_shape)
            if validate_bbox(bbox, image_shape):
                validated_details['container_number_bbox'] = bbox
            else:
                validated_details['container_number_bbox'] = None
        
        # Validate data plate info bbox
        if 'data_plate_info' in validated_details and isinstance(validated_details['data_plate_info'], dict):
            if validated_details['data_plate_info'].get('bbox'):
                bbox = normalize_bbox(validated_details['data_plate_info']['bbox'], image_shape)
                if validate_bbox(bbox, image_shape):
                    validated_details['data_plate_info']['bbox'] = bbox
                else:
                    validated_details['data_plate_info']['bbox'] = None
        
        # Validate location codes bboxes
        if 'location_codes' in validated_details and validated_details['location_codes']:
            valid_codes = []
            for loc_code in validated_details['location_codes']:
                if loc_code.get('bbox'):
                    bbox = normalize_bbox(loc_code['bbox'], image_shape)
                    if validate_bbox(bbox, image_shape):
                        loc_code['bbox'] = bbox
                        valid_codes.append(loc_code)
                    else:
                        # Keep code but remove invalid bbox
                        loc_code_copy = loc_code.copy()
                        loc_code_copy.pop('bbox', None)
                        valid_codes.append(loc_code_copy)
                else:
                    valid_codes.append(loc_code)
            validated_details['location_codes'] = valid_codes
        
        return validated_details
    
    def _calculate_class_counts(self, damages: list) -> dict:
        """Calculate per-class damage counts for YOLO summary"""
        class_counts = {
            "0_dent_deformation": 0,
            "1_corrosion_rust": 0,
            "2_frame_bend": 0,
            "3_door_hardware_damage": 0,
            "4_hole_tear_crack": 0
        }
        
        # Map damage types to class IDs
        type_to_class = {
            'dent': 0,
            'dent_deformation': 0,
            'deformation': 0,
            'rust': 1,
            'corrosion': 1,
            'corrosion_rust': 1,
            'bent_frame': 2,
            'frame_bend': 2,
            'door': 3,
            'door_hardware_damage': 3,
            'hole': 4,
            'crack': 4,
            'hole_tear_crack': 4,
            'tear': 4
        }
        
        for damage in damages:
            damage_type = damage.get('type', 'unknown').lower()
            class_id = damage.get('class_id')
            
            # Try to get class_id from damage, or map from type
            if class_id is None:
                class_id = type_to_class.get(damage_type, 0)
            
            # Get class name
            class_names = {
                0: "0_dent_deformation",
                1: "1_corrosion_rust",
                2: "2_frame_bend",
                3: "3_door_hardware_damage",
                4: "4_hole_tear_crack"
            }
            
            class_key = class_names.get(class_id, "0_dent_deformation")
            class_counts[class_key] = class_counts.get(class_key, 0) + 1
        
        return class_counts
    
    def _add_location_codes_to_damages(self, damages: list, container_details: dict, image_shape: tuple = None) -> list:
        """Add ISO-style location_code to damage entries.
        If a damage already has location_code, keep it; otherwise derive deterministically from bbox.
        """
        damages_with_codes = []
        for damage in damages:
            damage_copy = damage.copy()

            existing = str(damage_copy.get('location_code') or '').strip()
            if existing and existing.upper() not in ('UNKNOWN', 'N/A', 'NA'):
                damage_copy['location_code'] = existing
                damage_copy['code_mode'] = 'ISO 6346'
                damages_with_codes.append(damage_copy)
                continue

            bbox = damage_copy.get('bbox')
            if isinstance(bbox, (list, tuple)) and len(bbox) >= 4:
                try:
                    x, y, w, h = int(bbox[0]), int(bbox[1]), int(bbox[2]), int(bbox[3])
                except Exception:
                    x = y = w = h = None
                if x is not None and w is not None and w > 0 and h > 0:
                    cx = int(x + (w / 2.0))
                    cy = int(y + (h / 2.0))
                    derived = self._derive_location_code(
                        bbox_center=(cx, cy),
                        image_shape=image_shape,
                        damage_type=str(damage_copy.get('type') or ''),
                        zone=str(damage_copy.get('zone') or ''),
                    )
                    if derived:
                        damage_copy['location_code'] = derived
                    else:
                        damage_copy['location_code'] = 'UNKNOWN'
                else:
                    damage_copy['location_code'] = 'UNKNOWN'
            else:
                damage_copy['location_code'] = 'UNKNOWN'

            damage_copy['code_mode'] = 'ISO 6346'
            
            damages_with_codes.append(damage_copy)
        
        return damages_with_codes

    def _load_location_code_catalog(self) -> set:
        catalog = set()
        xlsx_path = Path('Location code (1).xlsx')
        if xlsx_path.exists():
            try:
                wb = load_workbook(str(xlsx_path), read_only=True, data_only=True)
                ws = wb.active
                for row in ws.iter_rows(min_row=2, max_col=ws.max_column, values_only=True):
                    val = row[-1] if row else None
                    if val is None:
                        continue
                    code = str(val).strip()
                    if code:
                        catalog.add(code)
            except Exception:
                catalog = set()

        if not catalog:
            txt_path = Path('configs') / 'iso6346_location_codes.txt'
            if txt_path.exists():
                try:
                    with open(txt_path, 'r', encoding='utf-8') as f:
                        for line in f:
                            code = str(line).strip()
                            if code:
                                catalog.add(code)
                except Exception:
                    pass
        return catalog

    def _build_location_code_index(self, catalog: set) -> dict:
        index = {}
        for code in sorted(catalog):
            c = str(code).strip()
            if len(c) < 3:
                continue
            prefix = c[:2]
            m = re.search(r'(\d+)', c)
            num = int(m.group(1)) if m else None
            index.setdefault(prefix, []).append((num, c))

        for prefix, items in index.items():
            items.sort(key=lambda t: (t[0] is None, t[0] if t[0] is not None else 10**9, t[1]))
            index[prefix] = items
        return index

    def _derive_location_code(self, bbox_center: tuple, image_shape: tuple = None, damage_type: str = '', zone: str = '') -> str:
        if not bbox_center or len(bbox_center) < 2:
            return ''

        cx, cy = int(bbox_center[0]), int(bbox_center[1])
        img_h = img_w = None
        if isinstance(image_shape, (list, tuple)) and len(image_shape) >= 2:
            img_h, img_w = int(image_shape[0]), int(image_shape[1])

        z = str(zone or '').strip().lower()
        t = str(damage_type or '').strip().lower()

        if 'door' in z or 'door' in t:
            prefix = 'DR' if (img_w and cx >= img_w / 2.0) else 'DL'
        elif 'roof' in z:
            prefix = 'TX'
        elif 'floor' in z:
            prefix = 'BX'
        elif 'rail' in z:
            prefix = 'RB' if (img_h and cy >= img_h * 0.66) else 'TR'
        else:
            if img_h:
                if cy < img_h * 0.33:
                    v = 'U'
                elif cy > img_h * 0.66:
                    v = 'D'
                else:
                    v = 'T'
            else:
                v = 'T'

            if img_w:
                if cx < img_w * 0.33:
                    h = 'L'
                elif cx > img_w * 0.66:
                    h = 'R'
                else:
                    h = 'X'
            else:
                h = 'X'

            prefix = f"{v}{h}"

        if not self._location_code_catalog:
            return ''

        bays = 12
        if img_w:
            try:
                bays = max(1, int(self.config.get('inspection_postprocess', {}).get('location_bays', 12)))
            except Exception:
                bays = 12
            idx = int(round((cx / float(max(1, img_w))) * (bays - 1))) + 1
        else:
            idx = 1
        idx = max(1, min(idx, bays))

        direct = f"{prefix}{idx:02d}"
        if direct in self._location_code_catalog:
            return direct

        items = self._location_code_index.get(prefix, [])
        numeric_items = [(n, c) for (n, c) in items if n is not None]
        if numeric_items:
            best = min(numeric_items, key=lambda t: (abs(int(t[0]) - idx), t[1]))
            return best[1]

        if items:
            return items[0][1]

        return ''